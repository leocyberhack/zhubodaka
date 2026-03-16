from collections import defaultdict
from io import BytesIO

from openpyxl import Workbook


HEADERS = ("日期", "直播开始时间", "直播结束时间", "直播账号")
COLUMN_WIDTHS = {"A": 18.89, "B": 16.55, "C": 14.22, "D": 13.78}


def build_monthly_schedule_workbook(entries, target_year, target_month):
    if not entries:
        raise ValueError("所选月份没有可导出的排班记录")

    grouped = defaultdict(list)
    for entry in entries:
        grouped[entry.anchor_name].append(entry)

    workbook = Workbook()
    workbook.remove(workbook.active)

    used_titles = set()
    for anchor_name in sorted(grouped):
        sheet = workbook.create_sheet(title=make_sheet_title(anchor_name, used_titles))
        used_titles.add(sheet.title)
        sheet.append(HEADERS)
        for column_name, width in COLUMN_WIDTHS.items():
            sheet.column_dimensions[column_name].width = width

        anchor_entries = sorted(
            grouped[anchor_name],
            key=lambda item: (item.live_date, item.start_time, item.end_time, item.live_account),
        )
        for row_index, entry in enumerate(anchor_entries, start=2):
            sheet.cell(row=row_index, column=1, value=entry.live_date)
            sheet.cell(row=row_index, column=2, value=entry.start_time)
            sheet.cell(row=row_index, column=3, value=entry.end_time)
            sheet.cell(row=row_index, column=4, value=entry.live_account)
            sheet[f"A{row_index}"].number_format = "yyyy/m/d"
            sheet[f"B{row_index}"].number_format = "h:mm"
            sheet[f"C{row_index}"].number_format = "h:mm"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"{target_year}年{target_month}月排班表.xlsx"
    return output, filename


def make_sheet_title(anchor_name, used_titles):
    base_title = (anchor_name or "未命名主播").strip() or "未命名主播"
    base_title = base_title.replace("/", "_").replace("\\", "_").replace("*", "_")
    base_title = base_title.replace("?", "_").replace(":", "_").replace("[", "_").replace("]", "_")
    base_title = base_title[:31] or "主播"

    if base_title not in used_titles:
        return base_title

    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = f"{base_title[:31 - len(suffix)]}{suffix}"
        if candidate not in used_titles:
            return candidate
        counter += 1
